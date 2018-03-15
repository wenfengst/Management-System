# -*- coding:utf-8 -*- ＃  
import web
import datetime
#import hashlib
import openpyxl
import StringIO
#import math
#db = web.database(dbn='mysql', db='appointment', user='root', pw='123456')
db = web.database(dbn='sqlite', db='E:/sqlite/apponitment.db')
#db = web.database(dbn='sqlite', db='E:/Management-System-master/Management-System-master/Appointment_system(10.5)/apponitment.db')


def login(username, password, usertype):
    '''登录验证'''
    users = db.select('user', where='username=$username AND password=$password AND usertype=$usertype', vars=locals())
    return users

def get_items(table):
    return db.select(table[0],order=table[1][-1]+' DESC')

def get_item(table,item_id):
    try:
        return db.select(table[0], where=table[1][0]+'=$item_id', vars=locals())[0]
    except IndexError:
        return None

def new_item(table,dic):
    a = dic.keys()
    s = 'db.insert(table[0]'
    for i in range(len(dic)):
        s = s+','+a[i]+'=dic[a['+str(i)+']]'    
    exec(s+','+table[1][-1]+'=datetime.datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\'))')

def del_item(table,item_id):
    db.query('pragma foreign_keys=ON')
    db.delete(table[0], where=table[1][0]+"=$item_id", vars=locals())

def update_item(table,item_id,dic):
    try:
        a = dic.keys()
        s = 'db.update(table[0], where=\''+table[1][0]+'=$item_id\', vars=locals()'
        for i in range(len(dic)):
            s = s+','+a[i]+'=dic[a['+str(i)+']]'
        exec(s+','+table[1][-1]+'=datetime.datetime.now().strftime(\'%Y-%m-%d %H:%M:%S\'))')
    except:
        return None

def insert_teacher_info(table,bank_num,table_item):
    tExist = db.select(table, where="bank_number=$bank_num", vars=locals())
    cFlag = True
    for tIter in tExist:        
        cFlag = False
        break
    if cFlag:
        db.multiple_insert(table,table_item)

def update_class_approve(cid,mid,approve):
    db.update('class',where='cid=$cid',vars=locals(),mid=mid,approve=approve,ctime=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
def get_projects(wh):
    return db.query('select project.*,user.full_name from project,user where project.oid=user.uid'+wh+' order by ptime DESC')
    
def get_classes(wh):
    return db.query('select cid,project.pid as pid,cname,begindate,begintime,enddate,endtime,classroom.place,classroom.classroom,\
    project.pname,project.snumber,teacher.tname,approve,u1.full_name manager,u2.full_name organizer,ctime \
    from class inner join project on class.pid=project.pid inner join teacher on class.tid=teacher.tid \
    inner join classroom on class.crid=classroom.crid left join user u1 on class.mid=u1.uid inner join user u2 on project.oid=u2.uid '+wh)

'''    
def get_classrooms(wh):
    return db.query('select cid,project.pid as pid,cname,begindate,begintime,enddate,endtime,classroom.place,classroom.classroom,\
    project.pname,project.snumber,teacher.tname,approve,u1.full_name manager,u2.full_name organizer,ctime \
    from classroom left join class on classroom.crid=class.crid left join project on class.pid=project.pid left join teacher on class.tid=teacher.tid \
    left join user u1 on class.mid=u1.uid left join user u2 on project.oid=u2.uid '+wh)
'''
    
def get_teacher_by_name(teachername):
    return db.select('teacher', where='tname=$teachername', vars=locals())

def get_teacher_out_by_name(teachername):
    return db.select('teacher_out', where='tname=$teachername', vars=locals())

def get_conflict_class(c):
    classid = c['cid']
    begin = c['begindate'].day+0.5*c['begintime']
    end = c['enddate'].day+0.5*c['endtime']
    year = str(c['begindate'])[0:4]
    month = str(c['begindate'])[5:7]
    place = c['place']
    classroom = c['classroom']
    wh = 'where (strftime(\'%d\',begindate)+0.5*begintime between '+str(begin)+' and '+str(end)\
    +' or strftime(\'%d\',enddate)+0.5*endtime between '+str(begin)+' and '+str(end)\
    +' or (strftime(\'%d\',begindate)+0.5*begintime<'+str(begin)+' and strftime(\'%d\',enddate)+0.5*endtime>'+str(end)+'))'\
    +' and strftime(\'%m\',begindate)=\''+month+'\' and strftime(\'%Y\',begindate)=\''+year+'\''\
    +' and classroom.place=\''+place+'\''+' and classroom.classroom=\''+classroom+'\''\
    +' and (approve!=\''+'不通过'.decode('utf-8')+'\' or (approve=\''+'不通过'.decode('utf-8')+'\' and cid='+str(classid)+')) order by begindate'
    classes = list(get_classes(wh))
    return classes

def get_xlsx_payment_in(n,t,p):
    f = openpyxl.load_workbook(u'static/xlsx/附件2：酬金审批表（2016年师资选聘）.xlsx')
    sheet1 = f[u'明细']
    sheet1.cell(row = 3 ,column = 7,value=t)
    sheet1.cell(row = 3 ,column = 2,value=p)
    ridx =5
    for index,item in enumerate(n):        
        sheet1.cell(row = index%11+ridx, column=1,value=item['name'])
        sheet1.cell(row = index%11+ridx, column=2,value=item['cname'])
        sheet1.cell(row = index%11+ridx, column=3,value=item['title'])
        sheet1.cell(row = index%11+ridx, column=4,value=item['hour_count'])
        sheet1.cell(row = index%11+ridx, column=5,value=item['projectname'])
        sheet1.cell(row = index%11+ridx, column=6,value=item['bank_num'])
        sheet1.cell(row = index%11+ridx, column=7,value=item['total'])
        #values=[{'tname':item['name'],'bank_number':item['bank_num'],'title':item['title'],'salary':int(item['total'])/int(item['hour_count']),'company':item['cname'],'bank_name':item['bank_name']}]
        #insert_teacher_info('teacher',item['bank_num'],values)
        if (index+1)%11==0:
            ridx+=22
            sheet1.cell(row = ridx-2 ,column = 7,value=t)
            sheet1.cell(row = ridx-2 ,column = 2,value=p)
    #print ridx
    ridx=((ridx-1)/22+1)*22+1
    while ridx<1000:

        for i in range(8):
            sheet1.cell(row=ridx,column=i+1,value="")
        ridx+=1
    sio=StringIO.StringIO()
    f.save(sio)
    return sio.getvalue()

def get_xlsx_payment_out(n,t,p):
    f = openpyxl.load_workbook(u'static/xlsx/表5：酬金发放审批表-模板.xlsx')
    sheet1 = f[u'局外']
    sheet1.cell(row = 6 ,column = 4,value=p)
    ridx =9
    for index,item in enumerate(n):

        sheet1.cell(row = index%15+ridx, column=1,value=item['name'])
        sheet1.cell(row = index%15+ridx, column=2,value=item['sex'])
        sheet1.cell(row = index%15+ridx, column=3,value=item['id'])
        sheet1.cell(row = index%15+ridx, column=4,value=item['bank_num'])
        sheet1.cell(row = index%15+ridx, column=5,value=item['title'])
        sheet1.cell(row = index%15+ridx, column=6,value=item['hour_count'])        
        moneyArr=get_tax(int(item['total']))

        sheet1.cell(row = index%15+ridx, column=7,value=str(moneyArr[0]))
        sheet1.cell(row = index%15+ridx, column=8,value=str(moneyArr[1]))        
        sheet1.cell(row = index%15+ridx, column=9,value=str(moneyArr[2])) 
        sheet1.cell(row = index%15+ridx, column=10,value=str(moneyArr[3]))               
        sheet1.cell(row = index%15+ridx, column=11,value=str(moneyArr[4]))
        sheet1.cell(row = index%15+ridx, column=12,value=str(moneyArr[5]))  
        sheet1.cell(row = index%15+ridx, column=13,value=str(item['total']))       

        #values=[{'tname':item['name'],'bank_number':item['bank_num'],'title':item['title'],'salary':int(item['total'])/int(item['hour_count']),'sex':item['sex'],'id':item['id'],'bank_name':item['bank_name']}]
        #insert_teacher_info('teacher_out',item['bank_num'],values)   
        if (index+1)%15==0:
            ridx+=26
            sheet1.cell(row = ridx-3 ,column = 7,value=t)
            sheet1.cell(row = ridx-3 ,column = 2,value=p)
  
    ridx=((ridx-1)/26+1)*26+1
    while ridx<1000:
        for i in range(13):
            sheet1.cell(row=ridx,column=i+1,value="")
        ridx+=1
    sio=StringIO.StringIO()
    f.save(sio)
    return sio.getvalue()


def get_tax(raw_num):
    tax_zengzhi=0
    tax_chengjian=0
    tax_jiaoyu=0
    tax_difang=0
    tax_person=0
    total=0
    money_without_tax=0
    if raw_num<=500:
        total=raw_num
    elif raw_num*1.032167<=800:
        total=round(raw_num*1.032167,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_chengjian=round(total-tax_zengzhi-raw_num,2)
    elif raw_num*1.03-164.8<=4000:
        total=round((raw_num*1.03-164.8)/0.7979,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_person=round((money_without_tax-800)*0.2,2)
        tax_chengjian=round(total-tax_zengzhi-tax_person-raw_num,2)
    elif raw_num*1.03/0.8379<=20000:
        total=round(raw_num*1.03/0.8379,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_person=round(money_without_tax*0.8*0.2,2)
        tax_chengjian=round(total-tax_zengzhi-tax_person-raw_num,2)
    elif (1.03*raw_num-2060)/0.7579<=50000:
        total=round((1.03*raw_num-2060)/0.7579,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_person=round(money_without_tax*0.8*0.3-2000,2)
        tax_chengjian=round(total-tax_zengzhi-tax_person-raw_num,2)
    elif (1.03*raw_num-7210)/0.6779<=100000:
        total=round((1.03*raw_num-7210)/0.6779,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_person=round(money_without_tax*0.8*0.4-7000,2)
        tax_chengjian=round(total-tax_zengzhi-tax_person-raw_num,2)
    else:
        total=round((1.03*raw_num-7210)/0.6764,2)
        money_without_tax=total/1.03
        tax_zengzhi=round(money_without_tax*0.03,2)
        tax_difang=round(tax_zengzhi*0.03,2)     
        tax_jiaoyu=round(tax_zengzhi*0.02,2)   
        tax_person=round(money_without_tax*0.8*0.4-7000,2)
        tax_chengjian=round(total-tax_zengzhi-tax_person-raw_num-tax_difang-tax_jiaoyu,2)        
    return [total,tax_zengzhi,tax_chengjian,tax_jiaoyu,tax_difang,tax_person]

def get_xlsx_daopan_out(n):
    f = openpyxl.load_workbook(u'static/xlsx/薪酬费用导盘表.xlsx')   
    sheet1 = f[u'原版'] 
    ridx =2
    gzps_b_name=sheet1.cell(row = 2, column=1).value
    gzps_b_num=sheet1.cell(row = 2, column=2).value
    gzps_c_name=sheet1.cell(row = 2, column=3).value

    total_payment = 0;
    for index,item in enumerate(n):

        sheet1.cell(row = index+ridx, column=1,value=gzps_b_name)
        sheet1.cell(row = index+ridx, column=2,value=gzps_b_num)
        sheet1.cell(row = index+ridx, column=3,value=gzps_c_name)
        sheet1.cell(row = index+ridx, column=4,value=item['bank_name'])
        sheet1.cell(row = index+ridx, column=5,value=u'广东')
        sheet1.cell(row = index+ridx, column=6,value=u'广州')
        sheet1.cell(row = index+ridx, column=7,value=item['bank_num'])
        sheet1.cell(row = index+ridx, column=8,value=item['name'])    
        moneyArr=get_tax(int(item['total']))
        
        sheet1.cell(row = index+ridx, column=9,value=moneyArr[0])
        total_payment += moneyArr[0]
        sheet1.cell(row = index+ridx, column=10,value=item['projectname'])
        #values=[{'tname':item['name'],'bank_number':item['bank_num'],'title':item['title'],'salary':int(item['total'])/int(item['hour_count']),'sex':item['sex'],'id':item['id'],'bank_name':item['bank_name']}]
        #insert_teacher_info('teacher_out',item['bank_num'],values) 
    sheet1.cell(row = index+ridx+1, column=8,value=u'合计')    
    sheet1.cell(row = index+ridx+1, column=9,value=total_payment)    
    sheet1.cell(row = index+ridx+3, column=2,value=u'部门主管')    
    sheet1.cell(row = index+ridx+3, column=7,value=u'制表人')    
        
    
    sio=StringIO.StringIO()
    f.save(sio)
    return sio.getvalue()

def get_xlsx_daopan(n):
    f = openpyxl.load_workbook(u'static/xlsx/薪酬费用导盘表.xlsx')   
    sheet1 = f[u'原版'] 
    ridx =2
    gzps_b_name=sheet1.cell(row = 2, column=1).value
    gzps_b_num=sheet1.cell(row = 2, column=2).value
    gzps_c_name=sheet1.cell(row = 2, column=3).value

    total_payment = 0;
    for index,item in enumerate(n):

        sheet1.cell(row = index+ridx, column=1,value=gzps_b_name)
        sheet1.cell(row = index+ridx, column=2,value=gzps_b_num)
        sheet1.cell(row = index+ridx, column=3,value=gzps_c_name)
        sheet1.cell(row = index+ridx, column=4,value=item['bank_name'])
        sheet1.cell(row = index+ridx, column=5,value=u'广东')
        sheet1.cell(row = index+ridx, column=6,value=u'广州')
        sheet1.cell(row = index+ridx, column=7,value=item['bank_num'])
        sheet1.cell(row = index+ridx, column=8,value=item['name'])        
        sheet1.cell(row = index+ridx, column=9,value=item['total'])
        total_payment += int(item['total'])
        #values=[{'tname':item['name'],'bank_number':item['bank_num'],'title':item['title'],'salary':int(item['total'])/int(item['hour_count']),'company':item['cname'],'bank_name':item['bank_name']}]
        #insert_teacher_info('teacher',item['bank_num'],values)
        sheet1.cell(row = index+ridx, column=10,value=item['projectname'])

    sheet1.cell(row = index+ridx+1, column=8,value=u'合计')    
    sheet1.cell(row = index+ridx+1, column=9,value=total_payment)    
    sheet1.cell(row = index+ridx+3, column=2,value=u'部门主管')    
    sheet1.cell(row = index+ridx+3, column=7,value=u'制表人')    
        
    
    sio=StringIO.StringIO()
    f.save(sio)
    return sio.getvalue()
#<td>{{itm.projectname}}</td>
#                        <td>{{itm.name}}</td>
#                        <td>{{itm.cname}}</td>
#                        <td>{{itm.bank_num}}</td>
#                        <td>{{itm.title}}</td>
#                        <td>{{itm.hour_count}}</td>
#                        <td>{{itm.total}}</td>


#导入sql
# insert into teacher(tname, company, salary, bank_number,bank_name)
#       select tname, company, salary, bank_number,bank_name from tmp;


# insert into teacher_out(tname, id,sex, salary, bank_number,bank_name)
        #select tname, id,sex, salary, bank_num,bank_name from `99`;
